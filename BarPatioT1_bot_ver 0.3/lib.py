import openpyxl as op
from config import bottle_list_path
import sqlite3
from menu import Time
from pprint import pprint 
import loging


wb = op.load_workbook(bottle_list_path)
sheet = wb.active

class BottleList:
    # функция получения значения из ячейки .xlsx файла.
    def value(row, column):
        value = sheet.cell(row = row, column = column).value
        return value
    
    # Функция получение всех значений в .xlsx файле в формате словаря.
    def get_list() -> dict:
        drink_list = {}
        for i in range(4, sheet.max_row + 1):
            name = BottleList.value(i, 3)
            sku = BottleList.value(i, 4)
            dealer  = BottleList.value(i, 5)
            size = BottleList.value(i, 6)
            unit = BottleList.value(i, 7)
            if not name: continue
            name_inf = [i, sku, dealer, size, unit]
            drink_list[name] = name_inf
        return drink_list
    # Функция поиска в словаре, по неполному названию.
    def search(name, mode = 'full_info'):
        name = str(name).lower()
        name = name.strip() 
        list = BottleList.get_list()
        position = {}
        position_key = []
        position_value = []
        for i in list:
            if name in str(i).lower():
                position_key.append(i)
                position_value = list.get(i)
        if len(position_key) > 1:
            position_key.insert(0, 'Найдены следующие позиции:')
            position_key.append('Уточните название...')
            return position_key
        elif len(position_key) == 1 and mode == 'only_names': 
            return position_key
        elif len(position_key) <= 0:
            position_key.insert(0, 'Ничего не найдено')
            return position_key
        elif len(position_key) == 1 and mode == 'full_info':
            key = str(position_key)
            key = key.replace(key[0], "")
            key = key.replace(key[-1], "")
            position[key] = position_value
            return position 

class Sql_libs: 
    def create_requires_lib(require):
        week = Time.get_week("today") +1
        date = Time.get_date("today")
        requirement_libs = sqlite3.connect("requirement_libs.db")
        cursor = requirement_libs.cursor()
        cursor.execute("""CREATE TABLE IF NOT EXISTS require_libs 
                        (
                        week integer primary key,
                        date text,
                        Requirements text
                        )""")
        try:
            cursor.execute("INSERT INTO require_libs VALUES ('%s', '%s', '%s')" % (week, date, require))
            requirement_libs.commit()
        except sqlite3.IntegrityError: return 'err: RETRY_CREATE_REQUIREMENT'
        except Exception as e:
            loging.err_archivist(__name__, e) 
        else: return "done"
        finally:
            cursor.close()
            requirement_libs.close()
    def get_create_requires_lib(week = 'this'):
        # Условия для определения нужной пользователю SQL-таблицы по запрашиваемой неделе.
        if week == 'this':
            week_num = Time.get_week("today")
        elif week == 'last':
            week_num = Time.get_week("today") -1
        elif week == 'next':
            week_num = Time.get_week("today") +1
        # Подключение к SQL бд
        requirement_libs = sqlite3.connect("requirement_libs.db")
        cursor = requirement_libs.cursor()
        cursor.execute(f"SELECT * FROM require_libs WHERE week = {week_num}")
        try:
            require = cursor.fetchone()[2]
        except TypeError: # Если таблица вернет пустой картеж или не вернет вовсе.
            if week == 'this': # Если запрос на эту неделю вернул пустую таблицу.
                return 'err 0.1'
            elif week == 'last': # Если запрос на прошлую неделю вернул пустую таблицу.
                return 'err 0'
        except Exception as e:
            loging.err_archivist(__name__, e) 
        else:
            raw_require = str(require).replace('(', '').replace(')', '')
            list_require = raw_require.split('\n')
            for i in list_require:
                if not i in require: require += str(i) + '\n'
            return require
        finally:
            cursor.close()
            requirement_libs.close()


